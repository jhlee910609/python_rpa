import datetime
import re
import ssl
from telnetlib import EC
from urllib.request import urlopen

import requests
from selenium import webdriver
import csv
from os import listdir, makedirs
from os.path import isdir
import os
import urllib
import time
from bs4 import BeautifulSoup
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook, Workbook
from selenium.webdriver.support.wait import WebDriverWait
from multiprocessing import Pool

wb = load_workbook("./partner_list.xlsx", data_only=True)
for name in wb.sheetnames:
    if name != "파트너_리스트":
        print(name)
        sheet = wb[name]
        print(sheet.max_row)
