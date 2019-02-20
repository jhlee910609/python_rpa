from os import listdir
from openpyxl import load_workbook, Workbook

path = '/Users/JunHee/Desktop/python_book/examples/06/'
files = listdir(path)
result_xlsx = Workbook()
result_sheet = result_xlsx.active

for myFile in files:
    # 확장자 예외 처리 코드
    if myFile[-4:] != 'xlsx':
        continue

    print("=========%s"%(myFile))
    tg_xlsx = load_workbook(path + myFile, True)
    tg_sheet = tg_xlsx.active

    for row in tg_sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)

        result_sheet.append(row_data)


result_xlsx.save(path + "result.xlsx")

xlsx = load_workbook(path+"result.xlsx", True)
result_sheet = xlsx.active