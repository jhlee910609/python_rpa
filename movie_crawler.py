import datetime
import re
import ssl
import time
from multiprocessing.pool import Pool
from openpyxl import load_workbook

from openpyxl import load_workbook
from selenium import webdriver

def find_movie_url():
    movie_urls = []
    for sheet_name in all_sheet_name:
        temp_sheet = load_wb[sheet_name]

        for temp in temp_sheet:
            print(temp)

    return movie_urls


def get_content(partner_name, url):
    driver.get(url)
    time.sleep(1)

    inner_info = driver.find_element_by_class_name("inner_info")
    title = inner_info.find_element_by_class_name("tit_view").text
    date = inner_info.find_element_by_class_name("rel_date").find_element_by_class_name('emph_number').text
    view_count = inner_info.find_element_by_class_name('rel_read').find_element_by_class_name(
        'emph_number').text
    comment_count = inner_info.find_element_by_class_name('useutil_btn').find_element_by_class_name(
        'comment_count').text

    print("date : %s | name : %s | title : %s | count : %s | comment_count : %s | url : %s") % (
        date, partner_name, title, view_count, comment_count, url)
    new_sheet = load_wb.create_sheet("%s_movie" % partner_name)
    new_sheet.append([date, partner_name, title, view_count, comment_count, url])
    load_wb.save("./partner_list_2.xlsx")

if __name__ == '__main__':
    driver = webdriver.Chrome("./chromedriver")
    load_wb = load_workbook("./partner_list_2.xlsx")
    all_sheet_name = load_wb.sheetnames
    partner_list = []

    pool = Pool(processes=4)
    pool.map(get_content, partner_list, )
    load_wb.save("./partner_list_2.xlsx")
