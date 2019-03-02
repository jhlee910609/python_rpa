import datetime
import re
import ssl
import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException

SCROLL_PAUSE_TIME = 2
FILE_NAME = "./partner_list_2.xlsx"

def move_to_scroll_down(last_height):
    while True:
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")

        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)

        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")

        # break condition
        if new_height == last_height:
            break
        last_height = new_height


load_wb = load_workbook(FILE_NAME, data_only=True)


def get_partner_list_from_xlsx():
    parter_list = {}
    load_ws = load_wb["파트너_선별"]
    for row in load_ws:
        parter_list[row[0].value] = row[1].value
    return parter_list


def remove_tag(text):
    return re.sub('<.+?>', '', text, 0, re.I | re.S)


def make_new_sheet(partner_name):
    partner_sheet = load_wb.create_sheet(partner_name)
    # date, editor's name, title, view count, reply
    partner_sheet['A1'] = '날짜'
    partner_sheet['B1'] = '파트너명'
    partner_sheet['C1'] = '제목'
    partner_sheet['D1'] = '조회수'
    partner_sheet['E1'] = '댓글수'
    partner_sheet['F1'] = '주소'
    partner_sheet['G1'] = datetime.datetime.now()

def set_headless_opt():
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1920x1080')
    options.add_argument("disable-gpu")
    return options


partner_list = get_partner_list_from_xlsx()
start_time = datetime.datetime.now()
ssl._create_default_https_context = ssl._create_unverified_context
driver = webdriver.Chrome("./chromedriver", options=set_headless_opt())
del partner_list['파트너명']

for partner_name in partner_list.keys():
    make_new_sheet(partner_name)
    driver.get(partner_list.get(partner_name))
    time.sleep(2)
    move_to_scroll_down(driver.execute_script("return document.body.scrollHeight"))

    parent = driver.find_element_by_class_name("list_1boon")
    content_list = parent.find_elements_by_tag_name("li")
    print(">>> 파트너명 : %s, 콘텐츠 갯수 : %s" % (partner_name, len(content_list)))
    content_urls = []

    for content in content_list:
        content_urls.append(content.find_element_by_class_name("link_1boon").get_attribute("href"))

    for url in content_urls:

        driver.get(url)
        driver.implicitly_wait(5)
        partner_sheet = load_wb[partner_name]

        try:
            inner_view = driver.find_element_by_class_name('inner_view')
            tit_view = inner_view.find_element_by_class_name("tit_view")
            rel_view = inner_view.find_element_by_class_name('rel_view')
            # info
            view_count = rel_view.find_element_by_class_name('rel_read').find_element_by_class_name(
                'emph_number').text
            date = rel_view.find_element_by_class_name('emph_number').text
            comment_count = inner_view.find_element_by_class_name('useutil_btn').find_element_by_class_name(
                'comment_count').text

            print("date : %s | name : %s | title : %s | count : %s | comment_count : %s | url : %s" % (
                date, partner_name, tit_view.text, view_count, comment_count, url))
            #
            # # date, editor's name, title, view count, reply
            partner_sheet.append([date, partner_name, tit_view.text, view_count, comment_count, url])
            # load_wb.save("./partner_list.xlsx")

        except NoSuchElementException as e:
            print(e)
            # inner_info = driver.find_element_by_class_name("inner_info")
            # title = inner_info.find_element_by_class_name("tit_view").text
            # date = inner_info.find_element_by_class_name("rel_date").find_element_by_class_name('emph_number').text
            # view_count = inner_info.find_element_by_class_name('rel_read').find_element_by_class_name(
            #     'emph_number').text
            # comment_count = inner_info.find_element_by_class_name('useutil_btn').find_element_by_class_name(
            #     'comment_count').text
            # print("%s, %s, %s, %s, %s, %s")%(date, partner_name, time, view_count, comment_count, url)
            partner_sheet.append(["", "", "", "", "", url])
            load_wb.save(FILE_NAME)
            pass

        except Exception as e:
            print(e)
            load_wb.save(FILE_NAME)
            print('>>>> 오류!')
            print("duration : %s" % str(datetime.datetime.now() - start_time))
            pass

    load_wb.save(FILE_NAME)
driver.quit()
print("duration : %s" % str(datetime.datetime.now() - start_time))
print('>>> 크롤링 성공!!!!!!!!!1')
